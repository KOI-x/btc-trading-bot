from pathlib import Path
import openpyxl

EXCEL_FILE = Path("bitcoin_prices.xlsx")

def inicializar_bd():
    """Crea el archivo de base de datos si no existe."""
    if not EXCEL_FILE.exists():
        try:
            wb = openpyxl.Workbook()
            wb.active.append(["Fecha", "Precio USD", "Variación %", "Desviación S2F %"])
            wb.save(EXCEL_FILE)
        except Exception as e:
            print(f"[ADVERTENCIA] No se pudo crear el archivo de datos: {e}")


def guardar_registro(fecha: str, precio: float, desviacion: float = 0.0) -> float:
    """Guarda un registro de fecha y precio en el Excel.

    Returns the percentage variation compared to the previous price."""
    if not EXCEL_FILE.exists():
        inicializar_bd()
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE)
    except Exception as e:
        print(f"[ADVERTENCIA] No se pudo abrir el archivo de datos: {e}")
        return 0.0

    sheet = wb.active

    # Calculate percentage change with respect to previous price
    variacion = 0.0
    if sheet.max_row > 1:
        prev_price = sheet.cell(row=sheet.max_row, column=2).value
        if isinstance(prev_price, (int, float)) and prev_price != 0:
            variacion = ((precio - prev_price) / prev_price) * 100

    sheet.append([fecha, precio, variacion, desviacion])
    try:
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"[ADVERTENCIA] No se pudo guardar el registro en el Excel: {e}")
    return variacion
